from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError as DRFValidationError
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from django.core.exceptions import ValidationError
from django.db import models
from io import BytesIO
from django.conf import settings
from openpyxl import load_workbook
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from urllib.parse import quote
from datetime import date
import logging

logger = logging.getLogger(__name__)

from .models import (
    Role, Permission, User,
    PassengerCar, NormsPassengerCars, PassengerCarWaybill,
    PassengerCarWaybillRecord, OdometerFuelPassengerCar,
    FireTruck, NormsFireTruck, FireTruckWaybill,
    FireTruckWaybillRecord, OdometerFuelFireTruck,

    OperatingHoursCars,
    NormsOperatingHoursPassengerCar,
    NormsOperatingHoursFireTruck,
    TechnicalMaintenance,
    NormsTechnicalMaintenance,
)

from .serializers import (
    RoleSerializer, PermissionSerializer, UserSerializer,
    PassengerCarSerializer, NormsPassengerCarsSerializer,
    PassengerCarWaybillSerializer, PassengerCarWaybillRecordSerializer,
    OdometerFuelPassengerCarSerializer,
    FireTruckSerializer, NormsFireTruckSerializer,
    FireTruckWaybillSerializer, FireTruckWaybillRecordSerializer,
    OdometerFuelFireTruckSerializer,

    OperatingHoursCarsSerializer,
    NormsOperatingHoursPassengerCarSerializer,
    NormsOperatingHoursFireTruckSerializer,
    TechnicalMaintenanceSerializer,
    NormsTechnicalMaintenanceSerializer,
)

from .permissions import (
    CanViewUsers, CanCreateUsers, CanUpdateUsers, CanDeleteUsers, CanViewDrivers,
    CanViewDriversReports, CanDownloadDriversReports,

    CanViewRoles, CanCreateRoles, CanUpdateRoles, CanDeleteRoles,
    CanViewPermissions, CanCreatePermissions, CanUpdatePermissions, CanDeletePermissions,

    CanViewPassengerCars, CanCreatePassengerCars,
    CanUpdatePassengerCars, CanDeletePassengerCars,

    CanViewPassengerCarNorms, CanCreatePassengerCarNorms,
    CanUpdatePassengerCarNorms, CanDeletePassengerCarNorms,

    CanViewPassengerCarWaybills, CanCreatePassengerCarWaybills,
    CanUpdatePassengerCarWaybills, CanDeletePassengerCarWaybills,
    CanDownloadPassengerCarWaybills,
    CanCreatePassengerCarWaybillRecord,
    CanUpdatePassengerCarWaybillRecord,
    CanDeletePassengerCarWaybillRecord,

    CanViewFireTrucks, CanCreateFireTrucks,
    CanUpdateFireTrucks, CanDeleteFireTrucks,

    CanViewFireTruckNorms, CanCreateFireTruckNorms,
    CanUpdateFireTruckNorms, CanDeleteFireTruckNorms,

    CanViewFireTruckWaybills, CanCreateFireTruckWaybills,
    CanUpdateFireTruckWaybills, CanDeleteFireTruckWaybills,
    CanDownloadFireTruckWaybills,
    CanCreateFireTruckWaybillRecord,
    CanUpdateFireTruckWaybillRecord,
    CanDeleteFireTruckWaybillRecord,

    CanCreateTechnicalMaintenance, CanDeleteTechnicalMaintenance,
    CanUpdateTechnicalMaintenance, CanViewTechnicalMaintenance,
    CanViewOperatingHours,
)

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
CENTER_FONT = Font(name='Times New Roman', size=11)
BOLD_FONT = Font(name='Times New Roman', size=11, bold=True)
THIN_BORDER = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )


class SoftDeleteModelViewSet(viewsets.ModelViewSet):
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ================= РОЛИ / ПРАВА / ПОЛЬЗОВАТЕЛИ =================

class RoleViewSet(SoftDeleteModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewRoles()]
        elif self.action == 'create':
            return base + [CanCreateRoles()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateRoles()]
        elif self.action == 'destroy':
            return base + [CanDeleteRoles()]
        return base + [CanViewRoles()]


class PermissionViewSet(SoftDeleteModelViewSet):
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer

    def get_queryset(self):
        queryset = Permission.objects.filter(deleted_at__isnull=True)
        role_id = self.request.query_params.get('role')
        if role_id:
            queryset = queryset.filter(role_id=role_id)
        return queryset

    def get_permissions(self):
        base = [IsAuthenticated()]

        if self.action == 'current':
            return base
        elif self.action in ['list', 'retrieve']:
            return base + [CanViewPermissions()]
        elif self.action == 'create':
            return base + [CanCreatePermissions()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePermissions()]
        elif self.action == 'destroy':
            return base + [CanDeletePermissions()]

        return base + [CanViewPermissions()]

    @action(detail=False, methods=['get'])
    def current(self, request):
        user = request.user

        if not user or not hasattr(user, 'role'):
            return Response(
                {'detail': 'User or user role not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            permission = Permission.objects.get(role=user.role, deleted_at__isnull=True)
            serializer = self.get_serializer(permission)
            return Response(serializer.data)
        except Permission.DoesNotExist:
            return Response(
                {'detail': 'No permissions found for user role'},
                status=status.HTTP_404_NOT_FOUND
            )


class UserViewSet(SoftDeleteModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewUsers()]
        elif self.action == 'create':
            return base + [CanCreateUsers()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateUsers()]
        elif self.action == 'destroy':
            return base + [CanDeleteUsers()]
        elif self.action == 'drivers':
            return base + [CanViewDrivers()]
        elif self.action == 'drivers_report':
            return base + [CanViewDriversReports()]
        elif self.action == 'drivers_report_excel':
            return base + [CanDownloadDriversReports()]
        return base + [CanViewUsers()]

    def get_queryset(self):
        role_id = self.request.query_params.get('role')
        if role_id:
            return self.queryset.filter(role_id=role_id, deleted_at__isnull=True)
        return self.queryset.filter(deleted_at__isnull=True)

    @action(detail=False, methods=['get'])
    def drivers(self, request):
        drivers = self.queryset.filter(role_id=3, deleted_at__isnull=True)
        serializer = self.get_serializer(drivers, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='drivers-report')
    def drivers_report(self, request):
        driver_id = request.query_params.get('driver')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not driver_id or not from_str or not to_str:
            return Response({"detail": "Параметры driver, from и to обязательны"}, status=400)

        try:
            driver_id = int(driver_id)
        except (TypeError, ValueError):
            return Response({"detail": "Неверный формат параметра driver"}, status=400)

        driver = User.objects.filter(id=driver_id, role_id=3, deleted_at__isnull=True).first()
        if not driver:
            return Response({"detail": "Водитель не найден или не имеет роль водитель"}, status=404)

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response({"detail": "Неверный формат дат"}, status=400)

        passenger_records = PassengerCarWaybillRecord.objects.filter(
            passenger_car_waybill__driver_id=driver_id,
            passenger_car_waybill__deleted_at__isnull=True,
            passenger_car_waybill__date__gte=from_date,
            passenger_car_waybill__date__lte=to_date,
        ).select_related('passenger_car_waybill__driver', 'passenger_car_waybill__car')

        fire_records = FireTruckWaybillRecord.objects.filter(
            fire_truck_waybill__driver_id=driver_id,
            fire_truck_waybill__deleted_at__isnull=True,
            fire_truck_waybill__date__gte=from_date,
            fire_truck_waybill__date__lte=to_date,
        ).select_related('fire_truck_waybill__driver', 'fire_truck_waybill__car')

        result = []

        for rec in passenger_records:
            d = rec.passenger_car_waybill.driver
            result.append({
                'date': rec.passenger_car_waybill.date,
                'driver': f'{d.surname} {d.name} {d.last_name}',
                'car_type': 'passenger',
                'car_number': rec.passenger_car_waybill.car.number,
                'route_or_target': rec.target,
                'distance': rec.distance_total_km,
                'fuel_used_fact': rec.fuel_used,
                'fuel_used_normal': rec.fuel_used_normal,
            })

        for rec in fire_records:
            d = rec.fire_truck_waybill.driver
            route = f"{rec.target} {rec.driving_route or ''}".strip()
            result.append({
                'date': rec.fire_truck_waybill.date,
                'driver': f'{d.surname} {d.name} {d.last_name}',
                'car_type': 'fire_truck',
                'car_number': rec.fire_truck_waybill.car.number,
                'route_or_target': route,
                'distance': rec.distance_km,
                'fuel_used_fact': rec.fuel_used,
                'fuel_used_normal': rec.fuel_used_normal,
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='drivers-report-excel')
    def drivers_report_excel(self, request):
        driver_id = request.query_params.get('driver')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not driver_id or not from_str or not to_str:
            return Response({"detail": "Параметры driver, from и to обязательны"}, status=400)

        try:
            driver_id = int(driver_id)
        except (TypeError, ValueError):
            return Response({"detail": "Неверный формат параметра driver"}, status=400)

        driver = User.objects.filter(id=driver_id, role_id=3, deleted_at__isnull=True).first()
        if not driver:
            return Response({"detail": "Водитель не найден или не имеет роль водитель"}, status=404)

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response({"detail": "Неверный формат дат"}, status=400)

        passenger_records = PassengerCarWaybillRecord.objects.filter(
            passenger_car_waybill__driver_id=driver_id,
            passenger_car_waybill__deleted_at__isnull=True,
            passenger_car_waybill__date__gte=from_date,
            passenger_car_waybill__date__lte=to_date,
        ).select_related('passenger_car_waybill__driver', 'passenger_car_waybill__car').order_by('passenger_car_waybill__date', 'id')

        fire_records = FireTruckWaybillRecord.objects.filter(
            fire_truck_waybill__driver_id=driver_id,
            fire_truck_waybill__deleted_at__isnull=True,
            fire_truck_waybill__date__gte=from_date,
            fire_truck_waybill__date__lte=to_date,
        ).select_related('fire_truck_waybill__driver', 'fire_truck_waybill__car').order_by('fire_truck_waybill__date', 'id')

        template_path = settings.BASE_DIR / 'report_templates' / 'driver.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active

        ws['D2'] = f"{driver.surname} {driver.name} {driver.last_name}"
        ws['I2'] = from_date.strftime('%d.%m.%Y')
        ws['M2'] = to_date.strftime('%d.%m.%Y')

        data_start_row = 7
        row_idx = data_start_row

        for rec in passenger_records:
            wb_obj = rec.passenger_car_waybill
            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=2, value=wb_obj.car.number)
            ws.cell(row=row_idx, column=3, value=float(rec.fuel_before_departure)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=4, value='Легковой автомобиль')
            ws.cell(row=row_idx, column=5, value=rec.departure_time.strftime('%H:%M') if rec.departure_time else None)
            ws.cell(row=row_idx, column=6, value=rec.arrival_time.strftime('%H:%M') if rec.arrival_time else None)
            ws.cell(row=row_idx, column=7, value=rec.odometer_before).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=8, value=float(rec.fuel_used)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=9, value=float(rec.fuel_used_normal)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=10, value=float(rec.fuel_refueled)).fill = GREEN_FILL
            ws.cell(row=row_idx, column=11, value=float(rec.fuel_on_return)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=12, value=rec.odometer_after).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=13, value=float(rec.fuel_used_normal - rec.fuel_used) if rec.fuel_used_normal > rec.fuel_used else None).fill = GREEN_FILL
            ws.cell(row=row_idx, column=14, value=float(rec.fuel_used - rec.fuel_used_normal) if rec.fuel_used_normal < rec.fuel_used else None).fill = RED_FILL
            row_idx += 1

        for rec in fire_records:
            wb_obj = rec.fire_truck_waybill
            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=2, value=wb_obj.car.number)
            ws.cell(row=row_idx, column=3, value=float(rec.fuel_before_departure)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=4, value='Пожарный автомобиль')
            ws.cell(row=row_idx, column=5, value=rec.departure_time.strftime('%H:%M') if rec.departure_time else None)
            ws.cell(row=row_idx, column=6, value=rec.arrival_time.strftime('%H:%M') if rec.arrival_time else None)
            ws.cell(row=row_idx, column=7, value=rec.odometer_before).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=8, value=float(rec.fuel_used)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=9, value=float(rec.fuel_used_normal)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=10, value=float(rec.fuel_refueled)).fill = GREEN_FILL
            ws.cell(row=row_idx, column=11, value=float(rec.fuel_on_return)).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=12, value=rec.odometer_after).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=13, value=float(rec.fuel_used_normal - rec.fuel_used) if rec.fuel_used_normal > rec.fuel_used else None).fill = GREEN_FILL
            ws.cell(row=row_idx, column=14, value=float(rec.fuel_used - rec.fuel_used_normal) if rec.fuel_used_normal < rec.fuel_used else None).fill = RED_FILL
            row_idx += 1

        for r in range(data_start_row, row_idx):
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                cell.border = THIN_BORDER
                cell.font = CENTER_FONT
                cell.alignment = CENTER_ALIGN

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Отчет по водителю {driver.surname}_{driver.name}_{driver.last_name} {from_date.strftime('%d.%m.%Y')}-{to_date.strftime('%d.%m.%Y')}.xlsx"
        quoted_filename = quote(filename)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
        return response


# ================= ЛЕГКОВЫЕ =================

class PassengerCarViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCar.objects.all()
    serializer_class = PassengerCarSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewPassengerCars()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCars()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCars()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCars()]
        return base
    
    def list(self, request, *args, **kwargs):
        """Переопределяем список для логирования"""
        logger.warning('\n' + '='*80)
        logger.warning('[PassengerCarViewSet.list] НАЧАЛО')
        logger.warning(f'  query_params: {dict(request.query_params)}')
        logger.warning('='*80)
        
        return super().list(request, *args, **kwargs)


class NormsPassengerCarsViewSet(SoftDeleteModelViewSet):
    queryset = NormsPassengerCars.objects.all()
    serializer_class = NormsPassengerCarsSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action == 'for_date':
            return base + [CanViewPassengerCarNorms()]
        elif self.action in ['list', 'retrieve']:
            return base + [CanViewPassengerCarNorms()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCarNorms()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCarNorms()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCarNorms()]
        return base

    @action(detail=False, methods=['get'], url_path='for-date')
    def for_date(self, request):
        car_id = request.query_params.get('car')
        season = request.query_params.get('season')
        date_str = request.query_params.get('date')

        if not car_id or not season or not date_str:
            return Response({"detail": "Параметры car, season и date обязательны"}, status=400)

        doc_date = parse_date(date_str)
        if not doc_date:
            return Response({"detail": "Неверный формат date, ожидается YYYY-MM-DD"}, status=400)

        norm = (
            NormsPassengerCars.objects
            .filter(car_id=car_id, season=season, date__lte=doc_date)
            .order_by('-date', '-id')
            .first()
        )
        if not norm:
            return Response({"detail": "Норма не найдена"}, status=404)

        serializer = self.get_serializer(norm)
        return Response(serializer.data)


class OdometerFuelPassengerCarViewSet(SoftDeleteModelViewSet):
    queryset = OdometerFuelPassengerCar.objects.all()
    serializer_class = OdometerFuelPassengerCarSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve', 'last_record']:
            return base + [CanViewPassengerCars()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCars()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCars()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCars()]
        return base

    @action(detail=False, methods=['get'], url_path='last')
    def last_record(self, request):
        car_id = request.query_params.get('car')
        if not car_id:
            return Response({"detail": "Параметр car обязателен"}, status=400)

        obj = (
            OdometerFuelPassengerCar.objects
            .filter(car_id=car_id)
            .order_by('-date', '-id')
            .first()
        )
        if not obj:
            return Response({"detail": "Записей не найдено"}, status=404)

        serializer = self.get_serializer(obj)
        return Response(serializer.data)


class PassengerCarWaybillViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCarWaybill.objects.all()
    serializer_class = PassengerCarWaybillSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewPassengerCarWaybills()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCarWaybills()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCarWaybills()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCarWaybills()]
        elif self.action == 'export_excel':
            return base + [CanDownloadPassengerCarWaybills()]
        return base

    def perform_update(self, serializer):
        """Валидация редактирования путевого листа - запретить изменение даты если он старше 7 дней"""
        waybill = self.get_object()
        
        # Проверяем что путевой лист еще editable
        if not waybill.is_editable():
            raise DRFValidationError(
                f"Невозможно редактировать путевой лист от {waybill.date}. "
                f"Редактирование разрешено только для путевых листов, созданных не более 7 дней назад."
            )
        
        # Проверяем что дата не была изменена (если путевой лист был editable и остается editable)
        if 'date' in serializer.validated_data:
            new_date = serializer.validated_data['date']
            if new_date != waybill.date:
                raise DRFValidationError(
                    "Редактирование даты путевого листа запрещено. "
                    "Дата путевого листа определяет доступность редактирования на 7 дней."
                )
        
        serializer.save()

    def perform_destroy(self, instance):
        """При удалении путевого листа - запустить cascade на первую запись следующего путевого листа"""
        waybill = instance
        car = waybill.car
        
        # Найти первую запись следующего путевого листа
        next_waybill = (
            PassengerCarWaybill.objects
            .filter(
                car=car,
                date__gt=waybill.date,
                deleted_at__isnull=True
            )
            .order_by('date', 'id')
            .first()
        )
        
        next_record = None
        if next_waybill:
            next_record = (
                PassengerCarWaybillRecord.objects
                .filter(passenger_car_waybill=next_waybill)
                .order_by('id')
                .first()
            )
        
        # Удаляем путевой лист
        super().perform_destroy(instance)
        
        # Если есть следующая запись - запускаем cascade
        if next_record:
            try:
                next_record.recalc_cascade()
            except Exception as e:
                logger.error(f'ОШИБКА при cascade после удаления путевого листа: {str(e)}')
                # Не прерываем удаление, просто логируем ошибку

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        car_id = request.query_params.get('car')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not car_id or not from_str or not to_str:
            return Response({"detail": "Параметры car, from и to обязательны"}, status=400)

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response({"detail": "Неверный формат дат, используйте YYYY-MM-DD"}, status=400)

        records = (
            PassengerCarWaybillRecord.objects
            .filter(
                passenger_car_waybill__car_id=car_id,
                passenger_car_waybill__deleted_at__isnull=True,
                passenger_car_waybill__date__gte=from_date,
                passenger_car_waybill__date__lte=to_date,
            )
            .select_related('passenger_car_waybill__driver', 'passenger_car_waybill__car')
            .order_by('passenger_car_waybill__date', 'id')
        )

        if not records.exists():
            return Response({"detail": "Записей за указанный период не найдено"}, status=404)

        car = records.first().passenger_car_waybill.car

        template_path = settings.BASE_DIR / 'report_templates' / 'passenger_car.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active

        ws['D2'] = car.number
        ws['I2'] = from_date.strftime('%d.%m.%Y')
        ws['N2'] = to_date.strftime('%d.%m.%Y')

        data_start_row = 7
        row_idx = data_start_row

        total_distance_city = 0
        total_distance_area = 0
        total_distance = 0
        total_fuel_used_city = 0
        total_fuel_used_area = 0
        total_fuel_used_fact = 0
        total_fuel_used_normal = 0
        total_fuel_refueled = 0
        total_savings = 0
        total_overrun = 0

        for rec in records:
            wb_obj = rec.passenger_car_waybill
            driver = wb_obj.driver

            savings = 0
            overrun = 0

            fio = f"{driver.surname} {driver.name[0]}. {driver.last_name[0]}."

            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=2, value=fio)

            ws.cell(row=row_idx, column=3, value=rec.fuel_before_departure).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=4, value=rec.odometer_before).fill = YELLOW_FILL

            ws.cell(row=row_idx, column=5, value=rec.distance_total_km)
            ws.cell(row=row_idx, column=6, value=rec.distance_city_km)
            ws.cell(row=row_idx, column=7, value=rec.distance_area_km)
            ws.cell(row=row_idx, column=8, value=rec.fuel_used_city)
            ws.cell(row=row_idx, column=9, value=rec.fuel_used_area)

            ws.cell(row=row_idx, column=10, value=rec.fuel_used_normal).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=11, value=rec.fuel_used).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=12, value=rec.fuel_refueled).fill = GREEN_FILL
            ws.cell(row=row_idx, column=13, value=rec.fuel_on_return).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=14, value=rec.odometer_after).fill = YELLOW_FILL

            savings = rec.fuel_used_normal - rec.fuel_used if rec.fuel_used_normal > rec.fuel_used else 0
            overrun = rec.fuel_used - rec.fuel_used_normal if rec.fuel_used_normal < rec.fuel_used else 0

            ws.cell(row=row_idx, column=15, value=float(savings)).fill = GREEN_FILL
            ws.cell(row=row_idx, column=16, value=float(overrun)).fill = RED_FILL

            total_distance_city += rec.distance_city_km
            total_distance_area += rec.distance_area_km
            total_distance += rec.distance_total_km
            total_fuel_used_city += rec.fuel_used_city
            total_fuel_used_area += rec.fuel_used_area
            total_fuel_used_fact += rec.fuel_used
            total_fuel_used_normal += rec.fuel_used_normal
            total_fuel_refueled += rec.fuel_refueled
            total_savings += savings
            total_overrun += overrun

            row_idx += 1

        ws.cell(row=row_idx, column=2, value="ИТОГО").fill = YELLOW_FILL
        ws.cell(row=row_idx, column=5, value=total_distance).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=6, value=total_distance_city).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=7, value=total_distance_area).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=8, value=float(total_fuel_used_city)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=9, value=float(total_fuel_used_area)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=10, value=float(total_fuel_used_normal)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=11, value=float(total_fuel_used_fact)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=12, value=float(total_fuel_refueled)).fill = GREEN_FILL
        ws.cell(row=row_idx, column=15, value=float(total_savings)).fill = GREEN_FILL
        ws.cell(row=row_idx, column=16, value=float(total_overrun)).fill = RED_FILL

        for r in range(data_start_row, row_idx + 1):
            for c in range(1, 17):
                cell = ws.cell(row=r, column=c)
                cell.border = THIN_BORDER
                cell.font = CENTER_FONT
                cell.alignment = CENTER_ALIGN

        for c in range(1, 17):
            ws.cell(row=row_idx, column=c).font = BOLD_FONT

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Путевые листы легкового автомобиля({car.number}) за период {from_date.strftime('%d.%m.%Y')}-{to_date.strftime('%d.%m.%Y')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        quoted_filename = quote(filename)
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
        return response


class PassengerCarWaybillRecordViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCarWaybillRecord.objects.select_related('passenger_car_waybill')
    serializer_class = PassengerCarWaybillRecordSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by passenger_car_waybill if provided in query params
        waybill_id = self.request.query_params.get('passenger_car_waybill')
        if waybill_id:
            queryset = queryset.filter(passenger_car_waybill_id=waybill_id)
        
        return queryset

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewPassengerCarWaybills()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCarWaybillRecord()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCarWaybillRecord()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCarWaybillRecord()]
        return base

    def perform_create(self, serializer):
        try:
            serializer.save()
        except ValidationError as e:
            # Преобразуем ValidationError в DRF format
            error_detail = e.message if hasattr(e, 'message') else str(e)
            raise DRFValidationError(error_detail)

    def perform_update(self, serializer):
        # Serializer уже блокирует расчетные поля если путевой лист старше 7 дней
        # Позволяем редактирование non-calculation полей в старых путевых листах
        record = self.get_object()
        
        # Сохраняем старые значения расчетных полей для проверки изменений
        old_fuel_refueled = record.fuel_refueled
        old_fuel_used = record.fuel_used
        old_odometer_after = record.odometer_after
        old_distance_city_km = record.distance_city_km
        old_distance_area_km = record.distance_area_km
        
        try:
            serializer.save()
            # Обновляем запись с актуальными данными из БД
            record.refresh_from_db()
            
            # Каскадный пересчет ТОЛЬКО если изменились расчетные поля
            calculation_fields_changed = (
                old_fuel_refueled != record.fuel_refueled or
                old_fuel_used != record.fuel_used or
                old_odometer_after != record.odometer_after or
                old_distance_city_km != record.distance_city_km or
                old_distance_area_km != record.distance_area_km
            )
            
            if calculation_fields_changed:
                record.recalc_cascade()
            
            # Пересчитываем totals путевого листа
            record.passenger_car_waybill.refresh_from_db()
            record.passenger_car_waybill.recalc_totals()
        except ValidationError as e:
            # Преобразуем ValidationError в DRF format
            error_detail = e.message if hasattr(e, 'message') else str(e)
            raise DRFValidationError(error_detail)

    def destroy(self, request, *args, **kwargs):
        # Проверка: запись должна быть в пределах 7 дней для удаления
        record = self.get_object()
        if not record.passenger_car_waybill.is_editable():
            raise DRFValidationError(
                f"Невозможно удалить запись из путевого листа от {record.passenger_car_waybill.date}. "
                f"Удаление разрешено только для путевых листов, созданных не более 7 дней назад."
            )
        
        # Найти следующую запись для cascade пересчета
        next_record = (
            PassengerCarWaybillRecord.objects
            .filter(
                passenger_car_waybill__car=record.passenger_car_waybill.car,
                passenger_car_waybill__deleted_at__isnull=True
            )
            .filter(
                models.Q(passenger_car_waybill__date__gt=record.passenger_car_waybill.date) |
                models.Q(passenger_car_waybill__date=record.passenger_car_waybill.date, id__gt=record.id)
            )
            .select_related('passenger_car_waybill')
            .order_by('passenger_car_waybill__date', 'id')
            .first()
        )
        
        # Удаляем запись
        response = super().destroy(request, *args, **kwargs)
        
        # Если есть следующая запись - запускаем cascade
        if next_record:
            try:
                next_record.recalc_cascade()
            except Exception as e:
                logger.error(f'❌ ОШИБКА при cascade после удаления записи: {str(e)}')
                # Не прерываем удаление, просто логируем ошибку
        
        return response


# ================= ПОЖАРНЫЕ =================

class FireTruckViewSet(SoftDeleteModelViewSet):
    queryset = FireTruck.objects.all()
    serializer_class = FireTruckSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewFireTrucks()]
        elif self.action == 'create':
            return base + [CanCreateFireTrucks()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTrucks()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTrucks()]
        return base
    
    def list(self, request, *args, **kwargs):
        """Переопределяем список для логирования"""
        logger.warning('\n' + '='*80)
        logger.warning('[FireTruckViewSet.list] НАЧАЛО')
        logger.warning(f'  query_params: {dict(request.query_params)}')
        logger.warning('='*80)
        
        return super().list(request, *args, **kwargs)


class NormsFireTruckViewSet(SoftDeleteModelViewSet):
    queryset = NormsFireTruck.objects.all()
    serializer_class = NormsFireTruckSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action == 'for_date':
            return base + [CanViewFireTruckNorms()]
        elif self.action in ['list', 'retrieve']:
            return base + [CanViewFireTruckNorms()]
        elif self.action == 'create':
            return base + [CanCreateFireTruckNorms()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTruckNorms()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTruckNorms()]
        return base

    @action(detail=False, methods=['get'], url_path='for-date')
    def for_date(self, request):
        car_id = request.query_params.get('car')
        season = request.query_params.get('season')
        date_str = request.query_params.get('date')

        if not car_id or not season or not date_str:
            return Response({"detail": "Параметры car, season и date обязательны"}, status=400)

        doc_date = parse_date(date_str)
        if not doc_date:
            return Response({"detail": "Неверный формат date, ожидается YYYY-MM-DD"}, status=400)

        norm = (
            NormsFireTruck.objects
            .filter(car_id=car_id, season=season, date__lte=doc_date)
            .order_by('-date', '-id')
            .first()
        )
        if not norm:
            return Response({"detail": "Норма не найдена"}, status=404)

        serializer = self.get_serializer(norm)
        return Response(serializer.data)


class OdometerFuelFireTruckViewSet(SoftDeleteModelViewSet):
    queryset = OdometerFuelFireTruck.objects.all()
    serializer_class = OdometerFuelFireTruckSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve', 'last_record']:
            return base + [CanViewFireTrucks()]
        elif self.action == 'create':
            return base + [CanCreateFireTrucks()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTrucks()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTrucks()]
        return base

    @action(detail=False, methods=['get'], url_path='last')
    def last_record(self, request):
        car_id = request.query_params.get('car')
        if not car_id:
            return Response({"detail": "Параметр car обязателен"}, status=400)

        obj = (
            OdometerFuelFireTruck.objects
            .filter(car_id=car_id)
            .order_by('-date', '-id')
            .first()
        )
        if not obj:
            return Response({"detail": "Записей не найдено"}, status=404)

        serializer = self.get_serializer(obj)
        return Response(serializer.data)


class FireTruckWaybillViewSet(SoftDeleteModelViewSet):
    queryset = FireTruckWaybill.objects.all()
    serializer_class = FireTruckWaybillSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewFireTruckWaybills()]
        elif self.action == 'create':
            return base + [CanCreateFireTruckWaybills()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTruckWaybills()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTruckWaybills()]
        elif self.action == 'export_excel':
            return base + [CanDownloadFireTruckWaybills()]
        return base

    def perform_update(self, serializer):
        """Валидация редактирования путевого листа - запретить изменение даты если он старше 7 дней"""
        waybill = self.get_object()
        
        # Проверяем что путевой лист еще editable
        if not waybill.is_editable():
            raise DRFValidationError(
                f"Невозможно редактировать путевой лист от {waybill.date}. "
                f"Редактирование разрешено только для путевых листов, созданных не более 7 дней назад."
            )
        
        # Проверяем что дата не была изменена (если путевой лист был editable и остается editable)
        if 'date' in serializer.validated_data:
            new_date = serializer.validated_data['date']
            if new_date != waybill.date:
                raise DRFValidationError(
                    "Редактирование даты путевого листа запрещено. "
                    "Дата путевого листа определяет доступность редактирования на 7 дней."
                )
        
        serializer.save()

    def perform_destroy(self, instance):
        """При удалении путевого листа - запустить cascade на первую запись следующего путевого листа"""
        waybill = instance
        car = waybill.car
        
        # Найти первую запись следующего путевого листа
        next_waybill = (
            FireTruckWaybill.objects
            .filter(
                car=car,
                date__gt=waybill.date,
                deleted_at__isnull=True
            )
            .order_by('date', 'id')
            .first()
        )
        
        next_record = None
        if next_waybill:
            next_record = (
                FireTruckWaybillRecord.objects
                .filter(fire_truck_waybill=next_waybill)
                .order_by('id')
                .first()
            )
        
        # Удаляем путевой лист
        super().perform_destroy(instance)
        
        # Если есть следующая запись - запускаем cascade
        if next_record:
            try:
                next_record.recalc_cascade()
            except Exception as e:
                logger.error(f'ОШИБКА при cascade после удаления путевого листа: {str(e)}')
                # Не прерываем удаление, просто логируем ошибку

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        car_id = request.query_params.get('car')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not car_id or not from_str or not to_str:
            return Response({"detail": "Параметры car, from и to обязательны"}, status=400)

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response({"detail": "Неверный формат дат, используйте YYYY-MM-DD"}, status=400)

        records = (
            FireTruckWaybillRecord.objects
            .filter(
                fire_truck_waybill__car_id=car_id,
                fire_truck_waybill__deleted_at__isnull=True,
                fire_truck_waybill__date__gte=from_date,
                fire_truck_waybill__date__lte=to_date,
            )
            .select_related('fire_truck_waybill__driver', 'fire_truck_waybill__car')
            .order_by('fire_truck_waybill__date', 'id')
        )

        if not records.exists():
            return Response({"detail": "Записей за указанный период не найдено"}, status=404)

        car = records.first().fire_truck_waybill.car

        template_path = settings.BASE_DIR / 'report_templates' / 'fire_truck.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active

        ws['D2'] = car.number
        ws['K2'] = from_date.strftime('%d.%m.%Y')
        ws['R2'] = to_date.strftime('%d.%m.%Y')

        data_start_row = 7
        row_idx = data_start_row

        total_distance_km = 0
        total_time_with_pump = 0
        total_time_without_pump = 0
        total_fuel_by_distance = 0
        total_fuel_with_pump = 0
        total_fuel_without_pump = 0
        total_fuel_normal = 0
        total_fuel_fact = 0
        total_fuel_refueled = 0
        total_savings = 0
        total_overrun = 0

        for rec in records:
            wb_obj = rec.fire_truck_waybill
            route = getattr(rec, 'driving_route', '') or ''
            name_place = (rec.target or '') + (f" {route}" if route else '')

            savings = 0
            overrun = 0

            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=2, value=name_place)

            ws.cell(row=row_idx, column=3, value=rec.fuel_before_departure).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=4, value=rec.departure_time.strftime('%H:%M'))
            ws.cell(row=row_idx, column=5, value=rec.arrival_time.strftime('%H:%M'))
            ws.cell(row=row_idx, column=6, value=rec.odometer_before).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=7, value=rec.distance_km)
            ws.cell(row=row_idx, column=8, value=rec.fuel_used_by_distance)
            ws.cell(row=row_idx, column=9, value=rec.time_with_pump)
            ws.cell(row=row_idx, column=10, value=rec.time_without_pump)
            ws.cell(row=row_idx, column=11, value=rec.fuel_used_with_pump)
            ws.cell(row=row_idx, column=12, value=rec.fuel_used_without_pump)

            ws.cell(row=row_idx, column=13, value=rec.fuel_used).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=14, value=rec.fuel_used_normal).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=15, value=rec.fuel_refueled).fill = GREEN_FILL
            ws.cell(row=row_idx, column=16, value=rec.fuel_on_return).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=17, value=rec.odometer_after).fill = YELLOW_FILL

            savings = rec.fuel_used_normal - rec.fuel_used if rec.fuel_used_normal > rec.fuel_used else 0
            overrun = rec.fuel_used - rec.fuel_used_normal if rec.fuel_used_normal < rec.fuel_used else 0

            ws.cell(row=row_idx, column=18, value=float(savings)).fill = GREEN_FILL
            ws.cell(row=row_idx, column=19, value=float(overrun)).fill = RED_FILL

            total_distance_km += rec.distance_km
            total_time_with_pump += rec.time_with_pump
            total_time_without_pump += rec.time_without_pump
            total_fuel_by_distance += rec.fuel_used_by_distance
            total_fuel_with_pump += rec.fuel_used_with_pump
            total_fuel_without_pump += rec.fuel_used_without_pump
            total_fuel_normal += rec.fuel_used_normal
            total_fuel_fact += rec.fuel_used
            total_fuel_refueled += rec.fuel_refueled
            total_savings += savings
            total_overrun += overrun

            row_idx += 1

        ws.cell(row=row_idx, column=2, value="ИТОГО").fill = YELLOW_FILL
        ws.cell(row=row_idx, column=7, value=total_distance_km).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=9, value=total_time_with_pump).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=10, value=total_time_without_pump).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=8, value=float(total_fuel_by_distance)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=11, value=float(total_fuel_with_pump)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=12, value=float(total_fuel_without_pump)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=13, value=float(total_fuel_fact)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=14, value=float(total_fuel_normal)).fill = YELLOW_FILL
        ws.cell(row=row_idx, column=15, value=float(total_fuel_refueled)).fill = GREEN_FILL
        ws.cell(row=row_idx, column=18, value=float(total_savings)).fill = GREEN_FILL
        ws.cell(row=row_idx, column=19, value=float(total_overrun)).fill = RED_FILL

        for r in range(data_start_row, row_idx + 1):
            for c in range(1, 20):
                cell = ws.cell(row=r, column=c)
                cell.border = THIN_BORDER
                cell.font = CENTER_FONT
                cell.alignment = CENTER_ALIGN

        for c in range(1, 20):
            ws.cell(row=row_idx, column=c).font = BOLD_FONT

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Путевые листы пожарного автомобиля({car.number}) за период {from_date.strftime('%d.%m.%Y')}-{to_date.strftime('%d.%m.%Y')}.xlsx"
        quoted_filename = quote(filename)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
        return response


class FireTruckWaybillRecordViewSet(SoftDeleteModelViewSet):
    queryset = FireTruckWaybillRecord.objects.select_related('fire_truck_waybill')
    serializer_class = FireTruckWaybillRecordSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by fire_truck_waybill if provided in query params
        waybill_id = self.request.query_params.get('fire_truck_waybill')
        if waybill_id:
            queryset = queryset.filter(fire_truck_waybill_id=waybill_id)
        
        return queryset

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewFireTruckWaybills()]
        elif self.action == 'create':
            return base + [CanCreateFireTruckWaybillRecord()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTruckWaybillRecord()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTruckWaybillRecord()]
        return base

    def perform_create(self, serializer):
        try:
            serializer.save()
        except ValidationError as e:
            # Преобразуем ValidationError в DRF format
            error_detail = e.message if hasattr(e, 'message') else str(e)
            raise DRFValidationError(error_detail)

    def perform_update(self, serializer):
        # Serializer уже блокирует расчетные поля если путевой лист старше 7 дней
        # Позволяем редактирование non-calculation полей в старых путевых листах
        record = self.get_object()
        
        # Сохраняем старые значения расчетных полей для проверки изменений
        old_fuel_refueled = record.fuel_refueled
        old_fuel_used = record.fuel_used
        old_odometer_after = record.odometer_after
        old_time_with_pump = record.time_with_pump
        old_time_without_pump = record.time_without_pump
        
        try:
            serializer.save()
            # Обновляем запись с актуальными данными из БД
            record.refresh_from_db()
            
            # Каскадный пересчет ТОЛЬКО если изменились расчетные поля
            calculation_fields_changed = (
                old_fuel_refueled != record.fuel_refueled or
                old_fuel_used != record.fuel_used or
                old_odometer_after != record.odometer_after or
                old_time_with_pump != record.time_with_pump or
                old_time_without_pump != record.time_without_pump
            )
            
            if calculation_fields_changed:
                record.recalc_cascade()
            
            # Пересчитываем totals путевого листа
            record.fire_truck_waybill.refresh_from_db()
            record.fire_truck_waybill.recalc_totals()
        except ValidationError as e:
            # Преобразуем ValidationError в DRF format
            error_detail = e.message if hasattr(e, 'message') else str(e)
            raise DRFValidationError(error_detail)

    def destroy(self, request, *args, **kwargs):
        # Проверка: запись должна быть в пределах 7 дней для удаления
        record = self.get_object()
        if not record.fire_truck_waybill.is_editable():
            raise DRFValidationError(
                f"Невозможно удалить запись из путевого листа от {record.fire_truck_waybill.date}. "
                f"Удаление разрешено только для путевых листов, созданных не более 7 дней назад."
            )
        
        # Найти следующую запись для cascade пересчета
        next_record = (
            FireTruckWaybillRecord.objects
            .filter(
                fire_truck_waybill__car=record.fire_truck_waybill.car,
                fire_truck_waybill__deleted_at__isnull=True
            )
            .filter(
                models.Q(fire_truck_waybill__date__gt=record.fire_truck_waybill.date) |
                models.Q(fire_truck_waybill__date=record.fire_truck_waybill.date, id__gt=record.id)
            )
            .select_related('fire_truck_waybill')
            .order_by('fire_truck_waybill__date', 'id')
            .first()
        )
        
        # Удаляем запись
        response = super().destroy(request, *args, **kwargs)
        
        # Если есть следующая запись - запускаем cascade
        if next_record:
            try:
                next_record.recalc_cascade()
            except Exception as e:
                logger.error(f'❌ ОШИБКА при cascade после удаления записи: {str(e)}')
                # Не прерываем удаление, просто логируем ошибку
        
        return response


# ================= МОТОЧАСЫ / ТО =================

class OperatingHoursCarsViewSet(SoftDeleteModelViewSet):
    queryset = OperatingHoursCars.objects.all().order_by('-date', '-id')
    serializer_class = OperatingHoursCarsSerializer
    http_method_names = ['get', 'head', 'options']

    def get_permissions(self):
        return [IsAuthenticated(), CanViewOperatingHours()]


class NormsOperatingHoursPassengerCarViewSet(SoftDeleteModelViewSet):
    queryset = NormsOperatingHoursPassengerCar.objects.all()
    serializer_class = NormsOperatingHoursPassengerCarSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewPassengerCarNorms()]
        elif self.action == 'create':
            return base + [CanCreatePassengerCarNorms()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdatePassengerCarNorms()]
        elif self.action == 'destroy':
            return base + [CanDeletePassengerCarNorms()]
        return base


class NormsOperatingHoursFireTruckViewSet(SoftDeleteModelViewSet):
    queryset = NormsOperatingHoursFireTruck.objects.all()
    serializer_class = NormsOperatingHoursFireTruckSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewFireTruckNorms()]
        elif self.action == 'create':
            return base + [CanCreateFireTruckNorms()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateFireTruckNorms()]
        elif self.action == 'destroy':
            return base + [CanDeleteFireTruckNorms()]
        return base


class NormsTechnicalMaintenanceViewSet(SoftDeleteModelViewSet):
    queryset = NormsTechnicalMaintenance.objects.all()
    serializer_class = NormsTechnicalMaintenanceSerializer

    def get_queryset(self):
        user = self.request.user
        role_perm = getattr(getattr(user, 'role', None), 'role', None)

        if not role_perm:
            return NormsTechnicalMaintenance.objects.none()

        can_passenger = getattr(role_perm, 'view_passenger_cars_norms', False)
        can_fire = getattr(role_perm, 'view_fire_truck_norms', False)

        qs = NormsTechnicalMaintenance.objects.all()

        if can_passenger and can_fire:
            # User can see both, but still apply query filters if provided
            pass
        elif can_passenger:
            qs = qs.filter(fire_truck__isnull=True)
        elif can_fire:
            qs = qs.filter(passenger_car__isnull=True)
        else:
            return qs.none()

        # Apply query parameter filters
        fire_truck_isnull = self.request.query_params.get('fire_truck__isnull')
        passenger_car_isnull = self.request.query_params.get('passenger_car__isnull')

        if fire_truck_isnull is not None:
            # fire_truck__isnull=false means show only fire trucks
            # fire_truck__isnull=true means show only passenger cars
            qs = qs.filter(fire_truck__isnull=fire_truck_isnull.lower() == 'true')

        if passenger_car_isnull is not None:
            # passenger_car__isnull=false means show only passenger cars
            # passenger_car__isnull=true means show only fire trucks
            qs = qs.filter(passenger_car__isnull=passenger_car_isnull.lower() == 'true')

        return qs

    def get_permissions(self):
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='remaining-hours')
    def remaining_hours(self, request):
        """
        GET /api/technical-maintenance-norms/remaining-hours/?passenger_car=<id>
        или
        GET /api/technical-maintenance-norms/remaining-hours/?fire_truck=<id>

        Возвращает, сколько осталось моточасов до каждого вида ТО.
        """
        passenger_car_id = request.query_params.get('passenger_car')
        fire_truck_id = request.query_params.get('fire_truck')
        date_str = request.query_params.get('date')

        if (not passenger_car_id and not fire_truck_id) or (passenger_car_id and fire_truck_id):
            return Response(
                {"detail": "Нужно указать либо passenger_car, либо fire_truck"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        calc_date = parse_date(date_str) if date_str else date.today()
        if date_str and not calc_date:
            return Response(
                {"detail": "Неверный формат date, используйте YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.user
        role_perm = getattr(getattr(user, 'role', None), 'role', None)
        if not role_perm:
            return Response({"detail": "Нет прав"}, status=status.HTTP_403_FORBIDDEN)

        if passenger_car_id and not getattr(role_perm, 'view_passenger_cars_norms', False):
            return Response({"detail": "Нет прав на просмотр норм легкового автомобиля"}, status=403)

        if fire_truck_id and not getattr(role_perm, 'view_fire_truck_norms', False):
            return Response({"detail": "Нет прав на просмотр норм пожарного автомобиля"}, status=403)

        result = []

        if passenger_car_id:
            last_hours = (
                OperatingHoursCars.objects
                .filter(passenger_car_id=passenger_car_id, date__lte=calc_date)
                .order_by('-date', '-id')
                .first()
            )
            current_hours = last_hours.operating_hours if last_hours else Decimal('0.000')

            norms = (
                NormsTechnicalMaintenance.objects
                .filter(passenger_car_id=passenger_car_id, fire_truck__isnull=True, date__lte=calc_date)
                .order_by('maintenance_type', '-date', '-id')
            )

            seen_types = set()
            for norm in norms:
                if norm.maintenance_type in seen_types:
                    continue
                seen_types.add(norm.maintenance_type)

                last_tm = (
                    TechnicalMaintenance.objects
                    .filter(passenger_car_id=passenger_car_id, maintenance_type=norm.maintenance_type, date__lte=calc_date)
                    .order_by('-date', '-id')
                    .first()
                )
                last_tm_hours = last_tm.operating_hours if last_tm else Decimal('0.000')
                spent_since = current_hours - last_tm_hours
                remaining = norm.norm - spent_since
                overdue = Decimal('0.000')
                if remaining < 0:
                    overdue = -remaining
                    remaining = Decimal('0.000')

                result.append({
                    'car_type': 'passenger',
                    'passenger_car': int(passenger_car_id),
                    'maintenance_type': norm.maintenance_type,
                    'maintenance_type_label': norm.get_maintenance_type_display(),
                    'norm': str(norm.norm),
                    'current_operating_hours': str(current_hours),
                    'last_maintenance_operating_hours': str(last_tm_hours),
                    'spent_since_last_maintenance': str(spent_since),
                    'remaining_hours': str(remaining),
                    'overdue_hours': str(overdue),
                    'is_due': overdue > 0 or remaining == 0,
                })

        if fire_truck_id:
            last_hours = (
                OperatingHoursCars.objects
                .filter(fire_truck_id=fire_truck_id, date__lte=calc_date)
                .order_by('-date', '-id')
                .first()
            )
            current_hours = last_hours.operating_hours if last_hours else Decimal('0.000')

            norms = (
                NormsTechnicalMaintenance.objects
                .filter(fire_truck_id=fire_truck_id, passenger_car__isnull=True, date__lte=calc_date)
                .order_by('maintenance_type', '-date', '-id')
            )

            seen_types = set()
            for norm in norms:
                if norm.maintenance_type in seen_types:
                    continue
                seen_types.add(norm.maintenance_type)

                last_tm = (
                    TechnicalMaintenance.objects
                    .filter(fire_truck_id=fire_truck_id, maintenance_type=norm.maintenance_type, date__lte=calc_date)
                    .order_by('-date', '-id')
                    .first()
                )
                last_tm_hours = last_tm.operating_hours if last_tm else Decimal('0.000')
                spent_since = current_hours - last_tm_hours
                remaining = norm.norm - spent_since
                overdue = Decimal('0.000')
                if remaining < 0:
                    overdue = -remaining
                    remaining = Decimal('0.000')

                result.append({
                    'car_type': 'fire_truck',
                    'fire_truck': int(fire_truck_id),
                    'maintenance_type': norm.maintenance_type,
                    'maintenance_type_label': norm.get_maintenance_type_display(),
                    'norm': str(norm.norm),
                    'current_operating_hours': str(current_hours),
                    'last_maintenance_operating_hours': str(last_tm_hours),
                    'spent_since_last_maintenance': str(spent_since),
                    'remaining_hours': str(remaining),
                    'overdue_hours': str(overdue),
                    'is_due': overdue > 0 or remaining == 0,
                })

        return Response(result)


class TechnicalMaintenanceViewSet(SoftDeleteModelViewSet):
    queryset = TechnicalMaintenance.objects.all()
    serializer_class = TechnicalMaintenanceSerializer

    def get_permissions(self):
        base = [IsAuthenticated()]
        if self.action in ['list', 'retrieve']:
            return base + [CanViewTechnicalMaintenance()]
        elif self.action == 'create':
            return base + [CanCreateTechnicalMaintenance()]
        elif self.action in ['update', 'partial_update']:
            return base + [CanUpdateTechnicalMaintenance()]
        elif self.action == 'destroy':
            return base + [CanDeleteTechnicalMaintenance()]
        elif self.action == 'perform_maintenance':
            return base + [CanCreateTechnicalMaintenance()]
        return base

    @action(detail=False, methods=['post'], url_path='perform')
    def perform_maintenance(self, request):
        """
        Провести техническое обслуживание машины.
        Создает запись ТО.
        
        Параметры:
        - car_id: ID легкового автомобиля (опционально)
        - truck_id: ID пожарного автомобиля (опционально)
        - maintenance_type: тип ТО (ТО-1, ТО-2, etc)
        - date: дата проведения ТО
        - spent: израсходовано топлива
        - received: получено топлива
        
        Примечание: operating_hours ВСЕГДА берутся из текущих значений в OperatingHoursCars.
                   Фронтенд не должен передавать это значение.
        """
        try:
            from decimal import Decimal
            from django.utils import timezone
            from datetime import datetime
            
            car_id = request.data.get('car_id')
            truck_id = request.data.get('truck_id')
            maintenance_type = request.data.get('maintenance_type')
            date = request.data.get('date')
            spent = request.data.get('spent', 0)
            received = request.data.get('received', 0)
            
            # Валидация обязательных полей
            if not maintenance_type:
                return Response({'error': 'maintenance_type обязателен'}, status=400)
            
            if not date:
                return Response({'error': 'date обязателена'}, status=400)
            
            if not car_id and not truck_id:
                return Response({'error': 'Необходимо указать car_id или truck_id'}, status=400)
            
            # Валидация на отрицательные значения
            try:
                spent_val = float(spent) if spent else 0
                received_val = float(received) if received else 0
                
                if spent_val < 0:
                    return Response({'error': 'spent не может быть отрицательным'}, status=400)
                if received_val < 0:
                    return Response({'error': 'received не может быть отрицательным'}, status=400)
            except (ValueError, TypeError):
                return Response({'error': 'Неверный формат числовых значений'}, status=400)
            
            # Получить машину
            passenger_car = None
            fire_truck = None
            if car_id:
                passenger_car = PassengerCar.objects.get(id=car_id)
            else:
                fire_truck = FireTruck.objects.get(id=truck_id)
            
            # operating_hours ВСЕГДА берутся из поля operating_hours машины (текущие значения)
            # Никогда не передаются с фронта и не берутся из OperatingHoursCars
            if passenger_car:
                operating_hours_val = float(passenger_car.operating_hours)
            else:
                operating_hours_val = float(fire_truck.operating_hours)
            
            print(f'[TechnicalMaintenance] 🔍 DEBUG perform_maintenance:')
            print(f'  - car_id: {car_id}, truck_id: {truck_id}')
            print(f'  - maintenance_type: {maintenance_type}')
            print(f'  - date: {date}')
            print(f'  - operating_hours_val (from car field): {operating_hours_val}')
            print(f'  - spent: {spent_val}, received: {received_val}')
            
            # Получить норму ТО (интервал)
            if passenger_car:
                norm_obj = NormsTechnicalMaintenance.objects.filter(
                    passenger_car=passenger_car,
                    maintenance_type=maintenance_type
                ).order_by('-date').first()
            else:
                norm_obj = NormsTechnicalMaintenance.objects.filter(
                    fire_truck=fire_truck,
                    maintenance_type=maintenance_type
                ).order_by('-date').first()
            
            if not norm_obj:
                return Response(
                    {'error': f'Норма ТО для {maintenance_type} не найдена'}, 
                    status=400
                )
            
            interval = float(norm_obj.norm)
            
            # Парсить дату
            try:
                maintenance_date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Неверный формат даты (используйте YYYY-MM-DD)'}, status=400)
            
            # Создать запись о проведении ТО
            maintenance = TechnicalMaintenance.objects.create(
                date=maintenance_date,
                car_type='passenger' if passenger_car else 'fire_truck',
                passenger_car=passenger_car,
                fire_truck=fire_truck,
                maintenance_type=maintenance_type,
                spent=Decimal(str(spent_val)),
                received=Decimal(str(received_val)),
                operating_hours=Decimal(str(operating_hours_val))
            )
            
            print(f'[TechnicalMaintenance] Создана запись TechnicalMaintenance:')
            print(f'  - id: {maintenance.id}')
            print(f'  - date: {maintenance.date}')
            print(f'  - maintenance_type: {maintenance.maintenance_type}')
            print(f'  - operating_hours: {maintenance.operating_hours}')
            
            serializer = TechnicalMaintenanceSerializer(maintenance)
            return Response({
                'success': True,
                'maintenance': serializer.data,
                'next_maintenance_at': float(operating_hours_val + interval)
            }, status=201)
        
        except PassengerCar.DoesNotExist:
            return Response({'error': 'Легковой автомобиль не найден'}, status=404)
        except FireTruck.DoesNotExist:
            return Response({'error': 'Пожарный автомобиль не найден'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=400)